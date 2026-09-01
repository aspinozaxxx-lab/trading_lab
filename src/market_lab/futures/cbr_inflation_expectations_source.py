"""Acquire release-specific CBR inflation-expectations source records."""

from __future__ import annotations

import argparse
import base64
import concurrent.futures
import gzip
import hashlib
import json
import math
import os
import re
import shutil
import tempfile
import zipfile
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, date, datetime
from datetime import time as wall_time
from decimal import ROUND_HALF_UP, Decimal
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Final, Protocol
from urllib.parse import urljoin, urlparse
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
import requests

from market_lab.io_utils import atomic_write_bytes, write_json

PROJECT_ROOT: Final[Path] = Path(__file__).resolve().parents[3]
ARCHIVE_URL: Final[str] = "https://www.cbr.ru/analytics/dkp/inflationary_expectations/"
SOURCE_START: Final[date] = date(2022, 1, 1)
SOURCE_END: Final[date] = date(2025, 12, 1)
PROTECTED_FROM: Final[pd.Timestamp] = pd.Timestamp("2026-01-01T00:00:00Z")
MOSCOW: Final[ZoneInfo] = ZoneInfo("Europe/Moscow")
DEFAULT_OUTPUT: Final[Path] = (
    PROJECT_ROOT
    / "data/processed/info_radar/"
    "cbr-inflation-expectations-release-pages-2022-2025-v1"
)
USER_AGENT: Final[str] = "market-lab-cbr-inflation-expectations/1.0 (causal research)"
DEFAULT_MAX_WORKERS: Final[int] = 8
MAX_HTML_BYTES: Final[int] = 2 * 1024 * 1024
MAX_PDF_BYTES: Final[int] = 24 * 1024 * 1024
MAX_XLSX_BYTES: Final[int] = 12 * 1024 * 1024
_ALLOWED_HOSTS: Final[frozenset[str]] = frozenset({"cbr.ru", "www.cbr.ru"})
_RELEASE_PATH: Final[re.Pattern[str]] = re.compile(
    r"^/analytics/dkp/inflationary_expectations/Infl_exp_"
    r"(?P<year>\d{2})-(?P<month>\d{2})/$",
    re.IGNORECASE,
)
_ATTACHMENT_PATH: Final[re.Pattern[str]] = re.compile(
    r"^/Collection/Collection/File/\d+/(?P<prefix>stat_)?Infl_exp_"
    r"(?P<year>\d{2})-(?P<month>\d{2})(?P<suffix>\.pdf|\.xlsx)$",
    re.IGNORECASE,
)
_SAFE_SNAPSHOT_ID: Final[re.Pattern[str]] = re.compile(
    r"^[A-Za-z0-9][A-Za-z0-9_.-]{0,127}$"
)
_RUSSIAN_MONTHS: Final[dict[str, int]] = {
    "января": 1,
    "февраля": 2,
    "марта": 3,
    "апреля": 4,
    "мая": 5,
    "июня": 6,
    "июля": 7,
    "августа": 8,
    "сентября": 9,
    "октября": 10,
    "ноября": 11,
    "декабря": 12,
}
_INFLATION_CHART_MARKER: Final[str] = (
    'id="GrafChart_ChartGroupModel_Charts_0__chart"'
)
_INFLATION_TABLE_LABEL: Final[str] = (
    "прямые оценки годовой инфляции: медианные значения"
)
_CSI_TABLE_LABEL: Final[str] = "индекс потребительских настроений (ипн)"
EXPECTED_RELEASE_MONTHS: Final[tuple[date, ...]] = tuple(
    date(year, month, 1)
    for year in range(SOURCE_START.year, SOURCE_END.year + 1)
    for month in range(1, 13)
    if SOURCE_START <= date(year, month, 1) <= SOURCE_END
)


class ResponseLike(Protocol):
    """Minimal requests-compatible response used by production and tests."""

    content: bytes
    headers: Mapping[str, str]

    def raise_for_status(self) -> None: ...


class SessionLike(Protocol):
    """Minimal requests-compatible session used by production and tests."""

    def get(
        self,
        url: str,
        *,
        headers: Mapping[str, str],
        timeout: float,
    ) -> ResponseLike: ...


@dataclass(frozen=True, slots=True)
class ReleaseLink:
    """One page and statistical workbook pair from the official archive."""

    release_month: date
    release_key: str
    page_url: str
    xlsx_url: str


@dataclass(frozen=True, slots=True)
class RequestRecord:
    """One raw official response retained in the external source archive."""

    kind: str
    identity: str
    url: str
    content: bytes
    headers: Mapping[str, str]


def sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(4 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _canonical_json(value: object) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")


def _official_url(value: str) -> str:
    absolute = urljoin(ARCHIVE_URL, value)
    parsed = urlparse(absolute)
    if parsed.scheme != "https" or parsed.hostname not in _ALLOWED_HOSTS:
        raise ValueError("CBR URL escaped the official HTTPS host")
    if parsed.username is not None or parsed.password is not None or parsed.port is not None:
        raise ValueError("CBR URL contains forbidden authority fields")
    return absolute


def _release_month_from_page(value: str) -> tuple[date, str] | None:
    candidate = urlparse(value)
    match = _RELEASE_PATH.fullmatch(candidate.path)
    if match is None:
        return None
    parsed = urlparse(_official_url(value))
    month = int(match.group("month"))
    year = 2000 + int(match.group("year"))
    try:
        release_month = date(year, month, 1)
    except ValueError as error:
        raise ValueError(f"invalid CBR inflation release path: {parsed.path}") from error
    return release_month, f"{year % 100:02d}-{month:02d}"


def _release_month_from_attachment(value: str, suffix: str) -> date | None:
    candidate = urlparse(value)
    match = _ATTACHMENT_PATH.fullmatch(candidate.path)
    if match is None or match.group("suffix").casefold() != suffix.casefold():
        return None
    parsed = urlparse(_official_url(value))
    month = int(match.group("month"))
    year = 2000 + int(match.group("year"))
    try:
        return date(year, month, 1)
    except ValueError as error:
        raise ValueError(f"invalid CBR inflation attachment path: {parsed.path}") from error


class _ArchiveParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.page_links: list[str] = []
        self.xlsx_links: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag != "a":
            return
        href = dict(attrs).get("href")
        if href is None:
            return
        if _release_month_from_page(href) is not None:
            self.page_links.append(_official_url(href))
        if _release_month_from_attachment(href, ".xlsx") is not None:
            self.xlsx_links.append(_official_url(href))


def discover_release_links(content: bytes) -> list[ReleaseLink]:
    """Discover exact page/XLSX pairs for the admitted release months."""
    if len(content) > MAX_HTML_BYTES:
        raise ValueError("CBR inflation archive page exceeds the HTML size limit")
    parser = _ArchiveParser()
    parser.feed(content.decode("utf-8-sig"))
    pages: dict[date, str] = {}
    workbooks: dict[date, str] = {}
    for url in sorted(set(parser.page_links)):
        parsed = _release_month_from_page(url)
        if parsed is None:
            continue
        release_month, _ = parsed
        if SOURCE_START <= release_month <= SOURCE_END:
            previous = pages.setdefault(release_month, url)
            if previous != url:
                raise ValueError(f"multiple CBR inflation pages for {release_month}")
    for url in sorted(set(parser.xlsx_links)):
        release_month = _release_month_from_attachment(url, ".xlsx")
        if release_month is not None and SOURCE_START <= release_month <= SOURCE_END:
            previous = workbooks.setdefault(release_month, url)
            if previous != url:
                raise ValueError(f"multiple CBR inflation workbooks for {release_month}")
    expected = set(EXPECTED_RELEASE_MONTHS)
    if set(pages) != expected or set(workbooks) != expected:
        raise ValueError(
            "CBR inflation release coverage drifted; "
            f"missing_pages={sorted(expected - set(pages))}, "
            f"missing_xlsx={sorted(expected - set(workbooks))}, "
            f"extra_pages={sorted(set(pages) - expected)}, "
            f"extra_xlsx={sorted(set(workbooks) - expected)}"
        )
    return [
        ReleaseLink(
            release_month=month,
            release_key=f"{month.year % 100:02d}-{month.month:02d}",
            page_url=pages[month],
            xlsx_url=workbooks[month],
        )
        for month in sorted(expected)
    ]


class _ReleasePageParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.links: list[str] = []
        self.visible_text: list[str] = []
        self.news_date_text: list[str] = []
        self._news_date_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        if self._news_date_depth:
            self._news_date_depth += 1
        classes = set((attributes.get("class") or "").split())
        if "news-info-line_date" in classes:
            if self._news_date_depth:
                raise ValueError("nested CBR publication-date containers")
            self._news_date_depth = 1
        if tag == "a" and attributes.get("href"):
            href = str(attributes["href"])
            if (
                _release_month_from_attachment(href, ".pdf") is not None
                or _release_month_from_attachment(href, ".xlsx") is not None
            ):
                self.links.append(_official_url(href))

    def handle_endtag(self, tag: str) -> None:
        del tag
        if self._news_date_depth:
            self._news_date_depth -= 1

    def handle_data(self, data: str) -> None:
        normalized = " ".join(data.split())
        if normalized:
            self.visible_text.append(normalized)
            if self._news_date_depth:
                self.news_date_text.append(normalized)


def _russian_page_date(value: str) -> date:
    normalized = " ".join(value.casefold().split())
    match = re.fullmatch(r"(\d{1,2})\s+([а-яё]+)\s+(\d{4})\s+года", normalized)
    if match is None or match.group(2) not in _RUSSIAN_MONTHS:
        raise ValueError(f"invalid CBR page date: {value!r}")
    return date(
        int(match.group(3)),
        _RUSSIAN_MONTHS[match.group(2)],
        int(match.group(1)),
    )


def conservative_available_at(publication_date: date, last_updated_date: date) -> pd.Timestamp:
    """Use Moscow day-end of the later printed page date."""
    if last_updated_date < publication_date:
        raise ValueError("CBR last-updated date precedes publication date")
    local = datetime.combine(
        max(publication_date, last_updated_date),
        wall_time(23, 59, 59),
        tzinfo=MOSCOW,
    )
    return pd.Timestamp(local.astimezone(UTC))


def _balanced_array(text: str, start: int) -> str:
    if start >= len(text) or text[start] != "[":
        raise ValueError("CBR chart array does not start with an opening bracket")
    depth = 0
    quoted = False
    escaped = False
    for index in range(start, len(text)):
        character = text[index]
        if quoted:
            if escaped:
                escaped = False
            elif character == "\\":
                escaped = True
            elif character == '"':
                quoted = False
            continue
        if character == '"':
            quoted = True
        elif character == "[":
            depth += 1
        elif character == "]":
            depth -= 1
            if depth == 0:
                return text[start : index + 1]
    raise ValueError("CBR chart contains an unbalanced JSON array")


def _array_after(text: str, key: str, start: int) -> tuple[object, int]:
    key_position = text.find(key, start)
    if key_position < 0:
        raise ValueError(f"CBR chart misses {key}")
    array_start = text.find("[", key_position + len(key))
    if array_start < 0:
        raise ValueError(f"CBR chart misses the array for {key}")
    encoded = _balanced_array(text, array_start)
    try:
        return json.loads(encoded), array_start + len(encoded)
    except json.JSONDecodeError as error:
        raise ValueError(f"CBR chart {key} is not valid JSON") from error


def _numeric_point(value: object, label: str) -> float:
    if isinstance(value, dict):
        value = value.get("y")
    if isinstance(value, bool):
        raise ValueError(f"CBR {label} chart point is boolean")
    try:
        result = float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"CBR {label} chart point is not numeric") from error
    if not math.isfinite(result):
        raise ValueError(f"CBR {label} chart point is not finite")
    return result


def _one_decimal(value: float) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP))


def _chart_components(page: str, release_month: date) -> dict[str, float]:
    marker_position = page.find(_INFLATION_CHART_MARKER)
    if marker_position < 0:
        raise ValueError("CBR release page misses the household inflation chart")
    categories_raw, categories_end = _array_after(page, '"categories":', marker_position)
    if not isinstance(categories_raw, list):
        raise ValueError("CBR inflation chart categories are not a list")
    categories: list[date] = []
    for group in categories_raw:
        if not isinstance(group, dict) or not isinstance(group.get("categories"), list):
            raise ValueError("CBR inflation chart category group is invalid")
        for raw_date in group["categories"]:
            try:
                categories.append(date.fromisoformat(raw_date))
            except (TypeError, ValueError) as error:
                raise ValueError("CBR inflation chart category is not an ISO date") from error
    if categories != sorted(set(categories)):
        raise ValueError("CBR inflation chart dates are not unique and chronological")
    try:
        release_index = categories.index(release_month)
    except ValueError as error:
        raise ValueError("CBR inflation chart lacks its release month") from error
    series_raw, _ = _array_after(page, '"series":', categories_end)
    if not isinstance(series_raw, list):
        raise ValueError("CBR inflation chart series are not a list")
    by_id = {
        item.get("id"): item
        for item in series_raw
        if isinstance(item, dict) and isinstance(item.get("id"), str)
    }
    required = {
        "s0": ("expected_inflation", "ожидаемая населением инфляция"),
        "s1": ("observed_inflation", "наблюдаемая населением инфляция"),
    }
    result: dict[str, float] = {}
    for series_id, (component, expected_name) in required.items():
        series = by_id.get(series_id)
        if series is None:
            raise ValueError(f"CBR inflation chart misses series {series_id}")
        if " ".join(str(series.get("name", "")).casefold().split()) != expected_name:
            raise ValueError(f"CBR inflation chart renamed series {series_id}")
        data = series.get("data")
        if not isinstance(data, list) or len(data) != len(categories):
            raise ValueError(f"CBR {component} data length differs from chart dates")
        if data[release_index] is None:
            raise ValueError(f"CBR {component} has no release-month point")
        if any(point is not None for point in data[release_index + 1 :]):
            raise ValueError(f"CBR {component} chart contains a future point")
        exact = _numeric_point(data[release_index], component)
        result[f"{component}_chart_exact"] = exact
        result[f"{component}_value"] = _one_decimal(exact)
    return result


def parse_release_page(
    content: bytes,
    *,
    release: ReleaseLink,
    retrieved_at_utc: str,
) -> dict[str, object]:
    """Parse availability and release-month inflation points from one page."""
    if len(content) > MAX_HTML_BYTES:
        raise ValueError("CBR inflation release page exceeds the HTML size limit")
    retrieved = pd.Timestamp(retrieved_at_utc)
    if retrieved.tzinfo is None:
        raise ValueError("CBR retrieval timestamp must be timezone-aware")
    page = content.decode("utf-8-sig")
    parser = _ReleasePageParser()
    parser.feed(page)
    publication_date = _russian_page_date(" ".join(parser.news_date_text))
    release_start = pd.Timestamp(release.release_month)
    publication = pd.Timestamp(publication_date)
    if publication < release_start or publication >= release_start + pd.DateOffset(months=2):
        raise ValueError("CBR publication date is implausible for its survey month")
    visible = " ".join(parser.visible_text)
    update_match = re.search(
        r"Последнее\s+обновление\s+страницы\s*:\s*(\d{2}\.\d{2}\.\d{4})",
        visible,
        re.IGNORECASE,
    )
    if update_match is None:
        raise ValueError("CBR inflation page misses its last-updated date")
    last_updated_date = datetime.strptime(update_match.group(1), "%d.%m.%Y").date()
    pdf_links = sorted(
        {
            link
            for link in parser.links
            if _release_month_from_attachment(link, ".pdf") == release.release_month
        }
    )
    if len(pdf_links) != 1:
        raise ValueError(f"unexpected CBR release PDF links: {pdf_links}")
    matching_xlsx = sorted(
        {
            link
            for link in parser.links
            if _release_month_from_attachment(link, ".xlsx") == release.release_month
        }
    )
    if matching_xlsx and matching_xlsx != [release.xlsx_url]:
        raise ValueError("CBR release page and archive disagree on the statistical workbook")
    return {
        "release_month": pd.Timestamp(release.release_month),
        "release_key": release.release_key,
        "publication_date": publication,
        "last_updated_date": pd.Timestamp(last_updated_date),
        "availability_date": pd.Timestamp(max(publication_date, last_updated_date)),
        "available_at": conservative_available_at(publication_date, last_updated_date),
        **_chart_components(page, release.release_month),
        "page_url": release.page_url,
        "pdf_url": pdf_links[0],
        "xlsx_url": release.xlsx_url,
        "retrieved_at_utc": retrieved,
        "release_specific_current_vintage": True,
        "modified_after_publication": last_updated_date > publication_date,
    }


def _normalize_label(value: object) -> str:
    return " ".join(str(value or "").casefold().split())


def _workbook_month(value: object) -> date:
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    if isinstance(value, str):
        parsed = date.fromisoformat(value[:10])
        return date(parsed.year, parsed.month, 1)
    raise ValueError(f"CBR workbook date cell is invalid: {value!r}")


def _find_label_row(rows: Sequence[Sequence[object]], label: str) -> int:
    matches = [
        row_number
        for row_number, row in enumerate(rows, start=1)
        if _normalize_label(row[0]) == label
    ]
    if len(matches) != 1:
        raise ValueError(f"CBR workbook label {label!r} matched rows {matches}")
    return matches[0]


def _find_series_title_row(rows: Sequence[Sequence[object]], label: str) -> int:
    candidates: list[int] = []
    for row_number, row in enumerate(rows, start=1):
        if _normalize_label(row[0]) != label:
            continue
        if row_number >= len(rows):
            continue
        date_count = 0
        for raw in rows[row_number][1:]:
            if raw is None:
                continue
            try:
                _workbook_month(raw)
            except (TypeError, ValueError):
                continue
            date_count += 1
        if date_count:
            candidates.append(row_number)
    if len(candidates) != 1:
        raise ValueError(f"CBR workbook series {label!r} matched rows {candidates}")
    return candidates[0]


def _workbook_endpoint(
    rows: Sequence[Sequence[object]],
    *,
    date_row: int,
    value_row: int,
    release_month: date,
    expected_value_label: str | None,
) -> float:
    if expected_value_label is not None:
        actual = _normalize_label(rows[value_row - 1][0])
        if actual != expected_value_label:
            raise ValueError(
                f"CBR workbook value row label drifted: {actual!r} != "
                f"{expected_value_label!r}"
            )
    points: list[tuple[date, float]] = []
    for raw_date, raw_value in zip(
        rows[date_row - 1][1:], rows[value_row - 1][1:], strict=True
    ):
        if raw_date is None and raw_value is None:
            continue
        month = _workbook_month(raw_date)
        if raw_value is None:
            continue
        value = _numeric_point(raw_value, "workbook")
        points.append((month, value))
    months = [month for month, _ in points]
    if not points or months != sorted(set(months)):
        raise ValueError("CBR workbook series is empty, duplicated, or nonchronological")
    if months[-1] != release_month:
        raise ValueError("CBR workbook series does not end on its release month")
    return points[-1][1]


def parse_statistics_workbook(content: bytes, *, release_month: date) -> dict[str, float]:
    """Parse exact household inflation and sentiment endpoints from a release XLSX."""
    if not content or len(content) > MAX_XLSX_BYTES:
        raise ValueError("CBR statistical workbook has an invalid size")
    if not content.startswith(b"PK") or not zipfile.is_zipfile(BytesIO(content)):
        raise ValueError("CBR statistical workbook is not an XLSX archive")
    try:
        workbook = openpyxl.load_workbook(
            BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
    except Exception as error:
        raise ValueError("CBR statistical workbook could not be opened") from error
    try:
        candidates = [
            worksheet
            for worksheet in workbook.worksheets
            if _normalize_label(worksheet.title) == "данные для графиков"
        ]
        if len(candidates) != 1:
            raise ValueError("CBR workbook has no unique data-for-charts sheet")
        worksheet = candidates[0]
        maximum_row = int(worksheet.max_row or 0)
        maximum_column = int(worksheet.max_column or 0)
        if not 1 <= maximum_row <= 2_048 or not 1 <= maximum_column <= 512:
            raise ValueError("CBR data-for-charts sheet dimensions are outside bounds")
        rows = list(
            worksheet.iter_rows(
                min_row=1,
                max_row=maximum_row,
                min_col=1,
                max_col=maximum_column,
                values_only=True,
            )
        )
        inflation_row = _find_label_row(rows, _INFLATION_TABLE_LABEL)
        observed = _workbook_endpoint(
            rows,
            date_row=inflation_row + 1,
            value_row=inflation_row + 2,
            release_month=release_month,
            expected_value_label="наблюдаемая инфляция",
        )
        expected = _workbook_endpoint(
            rows,
            date_row=inflation_row + 1,
            value_row=inflation_row + 3,
            release_month=release_month,
            expected_value_label="ожидаемая инфляция",
        )
        csi_row = _find_series_title_row(rows, _CSI_TABLE_LABEL)
        sentiment = _workbook_endpoint(
            rows,
            date_row=csi_row + 1,
            value_row=csi_row + 2,
            release_month=release_month,
            expected_value_label=None,
        )
        if not 0.0 <= expected <= 100.0 or not 0.0 <= observed <= 100.0:
            raise ValueError("CBR household inflation endpoint is outside a sanity range")
        if not 0.0 <= sentiment <= 200.0:
            raise ValueError("CBR consumer sentiment endpoint is outside a sanity range")
        return {
            "expected_inflation_exact": expected,
            "observed_inflation_exact": observed,
            "consumer_sentiment_index_exact": sentiment,
        }
    finally:
        workbook.close()


def _request_record(
    kind: str,
    identity: str,
    url: str,
    *,
    session: SessionLike,
) -> RequestRecord:
    official = _official_url(url)
    response = session.get(
        official,
        headers={"User-Agent": USER_AGENT, "Accept": "*/*"},
        timeout=60.0,
    )
    response.raise_for_status()
    content = response.content
    limit = {
        "pdf": MAX_PDF_BYTES,
        "xlsx": MAX_XLSX_BYTES,
    }.get(kind, MAX_HTML_BYTES)
    if not content or len(content) > limit:
        raise ValueError(f"CBR {kind} response has an invalid size")
    if kind == "pdf" and not content.startswith(b"%PDF-"):
        raise ValueError("CBR report attachment is not a PDF")
    if kind == "xlsx" and (
        not content.startswith(b"PK") or not zipfile.is_zipfile(BytesIO(content))
    ):
        raise ValueError("CBR statistical attachment is not XLSX")
    return RequestRecord(kind, identity, official, content, dict(response.headers))


def _fetch_many(
    requests_to_make: Sequence[tuple[str, str, str]],
    *,
    session: SessionLike,
    max_workers: int,
) -> list[RequestRecord]:
    def fetch(item: tuple[str, str, str]) -> RequestRecord:
        kind, identity, url = item
        try:
            return _request_record(kind, identity, url, session=session)
        except Exception as error:
            raise RuntimeError(f"CBR request failed for {kind}:{identity}:{url}") from error

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        return list(executor.map(fetch, requests_to_make))


def _atomic_parquet(path: Path, frame: pd.DataFrame) -> None:
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        frame.to_parquet(temporary, index=False, compression="zstd")
        temporary.replace(path)
    finally:
        temporary.unlink(missing_ok=True)


def _raw_record(record: RequestRecord) -> bytes:
    selected_headers = {
        name: record.headers.get(name)
        for name in ("Last-Modified", "Content-Type", "ETag")
        if record.headers.get(name) is not None
    }
    return json.dumps(
        {
            "kind": record.kind,
            "identity": record.identity,
            "url": record.url,
            "headers": selected_headers,
            "bytes": len(record.content),
            "sha256": sha256_bytes(record.content),
            "content_encoding": "base64",
            "content": base64.b64encode(record.content).decode("ascii"),
        },
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")


def download_cbr_inflation_expectations(
    output_directory: Path = DEFAULT_OUTPUT,
    *,
    session: SessionLike | None = None,
    max_workers: int = DEFAULT_MAX_WORKERS,
    fetched_at_utc: str | None = None,
    minimum_releases: int = 40,
) -> Path:
    """Write one immutable target-free source bundle; never overwrite it."""
    final = output_directory.resolve()
    if final.exists():
        raise FileExistsError(f"CBR inflation-expectations output already exists: {final}")
    if not _SAFE_SNAPSHOT_ID.fullmatch(final.name):
        raise ValueError("unsafe CBR inflation-expectations snapshot directory name")
    if max_workers < 1:
        raise ValueError("max_workers must be positive")
    fetched_at = fetched_at_utc or datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
    client = session or requests.Session()
    archive_initial = _request_record("archive", "initial", ARCHIVE_URL, session=client)
    releases = discover_release_links(archive_initial.content)
    page_records = _fetch_many(
        [("page", item.release_key, item.page_url) for item in releases],
        session=client,
        max_workers=max_workers,
    )
    xlsx_records = _fetch_many(
        [("xlsx", item.release_key, item.xlsx_url) for item in releases],
        session=client,
        max_workers=max_workers,
    )
    rows: list[dict[str, object]] = []
    for release, page_record, xlsx_record in zip(
        releases, page_records, xlsx_records, strict=True
    ):
        try:
            row = parse_release_page(
                page_record.content,
                release=release,
                retrieved_at_utc=fetched_at,
            )
            workbook_values = parse_statistics_workbook(
                xlsx_record.content, release_month=release.release_month
            )
        except ValueError as error:
            raise ValueError(
                f"CBR inflation release {release.release_key} could not be parsed"
            ) from error
        exact_matches: list[bool] = []
        for component in ("expected_inflation", "observed_inflation"):
            page_exact = float(row[f"{component}_chart_exact"])
            workbook_exact = float(workbook_values[f"{component}_exact"])
            if (
                abs(page_exact - workbook_exact) > 0.051
                or not math.isclose(
                    float(row[f"{component}_value"]),
                    _one_decimal(workbook_exact),
                    rel_tol=0.0,
                    abs_tol=1e-12,
                )
            ):
                raise ValueError(
                    f"CBR inflation page/XLSX mismatch for {release.release_key}:{component}"
                )
            exact_matches.append(
                math.isclose(page_exact, workbook_exact, rel_tol=0.0, abs_tol=1e-9)
            )
        rows.append(
            {
                **row,
                **workbook_values,
                "page_xlsx_display_match": True,
                "page_xlsx_exact_match": all(exact_matches),
            }
        )
    if len(rows) < minimum_releases:
        raise ValueError("CBR inflation release coverage is unexpectedly small")
    pdf_records = _fetch_many(
        [
            ("pdf", release.release_key, str(row["pdf_url"]))
            for release, row in zip(releases, rows, strict=True)
        ],
        session=client,
        max_workers=max_workers,
    )
    archive_final = _request_record("archive", "final", ARCHIVE_URL, session=client)
    if discover_release_links(archive_final.content) != releases:
        raise ValueError("CBR inflation archive changed during collection")
    frame = pd.DataFrame(rows).sort_values("release_month", ignore_index=True)
    if frame["release_month"].duplicated().any():
        raise ValueError("CBR inflation releases contain duplicate months")
    if set(frame["release_month"].dt.date) != set(EXPECTED_RELEASE_MONTHS):
        raise ValueError("CBR processed inflation coverage differs from the archive")
    if frame["available_at"].ge(PROTECTED_FROM).any():
        raise ValueError("CBR inflation release availability crossed the protected boundary")
    value_columns = [
        "expected_inflation_exact",
        "observed_inflation_exact",
        "consumer_sentiment_index_exact",
    ]
    if frame[value_columns].isna().any().any():
        raise ValueError("CBR inflation source contains missing strategy values")
    page_by_key = {record.identity: record for record in page_records}
    xlsx_by_key = {record.identity: record for record in xlsx_records}
    pdf_by_key = {record.identity: record for record in pdf_records}
    expected_keys = {release.release_key for release in releases}
    if set(page_by_key) != expected_keys or set(xlsx_by_key) != expected_keys:
        raise ValueError("CBR inflation page/XLSX response identities are incomplete")
    if set(pdf_by_key) != expected_keys:
        raise ValueError("CBR inflation PDF response identities are incomplete")
    coverage = pd.DataFrame(
        [
            {
                "release_month": row["release_month"],
                "publication_date": row["publication_date"],
                "last_updated_date": row["last_updated_date"],
                "available_at": row["available_at"],
                "release_key": release.release_key,
                "page_url": release.page_url,
                "page_bytes": len(page_by_key[release.release_key].content),
                "page_sha256": sha256_bytes(page_by_key[release.release_key].content),
                "pdf_url": row["pdf_url"],
                "pdf_bytes": len(pdf_by_key[release.release_key].content),
                "pdf_sha256": sha256_bytes(pdf_by_key[release.release_key].content),
                "xlsx_url": release.xlsx_url,
                "xlsx_bytes": len(xlsx_by_key[release.release_key].content),
                "xlsx_sha256": sha256_bytes(xlsx_by_key[release.release_key].content),
                "modified_after_publication": row["modified_after_publication"],
                "page_xlsx_display_match": row["page_xlsx_display_match"],
                "page_xlsx_exact_match": row["page_xlsx_exact_match"],
            }
            for release, row in zip(releases, rows, strict=True)
        ]
    ).sort_values("release_month", ignore_index=True)
    year_counts = Counter(frame["release_month"].dt.year.astype(int))
    expected_delta = frame["expected_inflation_exact"].diff().dropna()
    sentiment_delta = frame["consumer_sentiment_index_exact"].diff().dropna()
    aligned_risk_on = expected_delta.lt(0.0) & sentiment_delta.gt(0.0)
    aligned_risk_off = expected_delta.gt(0.0) & sentiment_delta.lt(0.0)
    availability_collisions = int(frame["available_at"].duplicated(keep=False).sum())
    availability_gaps = frame["available_at"].sort_values().diff().dt.total_seconds() / 86_400
    release_lag = (
        frame["publication_date"] - frame["release_month"]
    ).dt.total_seconds() / 86_400
    final.parent.mkdir(parents=True, exist_ok=True)
    temporary = Path(tempfile.mkdtemp(prefix=f".{final.name}.", dir=final.parent))
    try:
        data_path = temporary / "cbr_inflation_expectations_releases.parquet"
        coverage_path = temporary / "coverage.parquet"
        raw_path = temporary / "official_cbr_inflation_expectations_responses.jsonl.gz"
        _atomic_parquet(data_path, frame)
        _atomic_parquet(coverage_path, coverage)
        raw_records = [
            archive_initial,
            *page_records,
            *xlsx_records,
            *pdf_records,
            archive_final,
        ]
        raw_lines = [_raw_record(record) for record in raw_records]
        atomic_write_bytes(
            raw_path,
            gzip.compress(b"\n".join(raw_lines) + b"\n", compresslevel=6, mtime=0),
        )
        manifest_core = {
            "schema_version": 1,
            "source_id": "official-cbr-inflation-expectations-releases-2022-2025-v1",
            "provider": "Bank of Russia",
            "source_name": "Inflation expectations and consumer sentiment releases",
            "source_url": ARCHIVE_URL,
            "fetched_at_utc": fetched_at,
            "request_count": len(raw_records),
            "request_bounds": {
                "release_month_from": SOURCE_START.isoformat(),
                "release_month_through": SOURCE_END.isoformat(),
                "protected_from_utc": PROTECTED_FROM.isoformat(),
            },
            "coverage": {
                "release_pages": len(frame),
                "release_pages_by_year": {
                    str(year): year_counts[year] for year in sorted(year_counts)
                },
                "minimum_release_month": frame["release_month"].min().date().isoformat(),
                "maximum_release_month": frame["release_month"].max().date().isoformat(),
                "minimum_available_at": frame["available_at"].min().isoformat(),
                "maximum_available_at": frame["available_at"].max().isoformat(),
                "modified_after_publication_count": int(
                    frame["modified_after_publication"].sum()
                ),
                "rows_in_availability_collisions": availability_collisions,
                "maximum_availability_gap_days": float(availability_gaps.max()),
                "minimum_release_lag_days": float(release_lag.min()),
                "maximum_release_lag_days": float(release_lag.max()),
                "sequential_expected_inflation_delta_counts": {
                    "positive": int(expected_delta.gt(0.0).sum()),
                    "negative": int(expected_delta.lt(0.0).sum()),
                    "zero": int(expected_delta.eq(0.0).sum()),
                },
                "sequential_consumer_sentiment_delta_counts": {
                    "positive": int(sentiment_delta.gt(0.0).sum()),
                    "negative": int(sentiment_delta.lt(0.0).sum()),
                    "zero": int(sentiment_delta.eq(0.0).sum()),
                },
                "aligned_confirmation_counts": {
                    "risk_on": int(aligned_risk_on.sum()),
                    "risk_off": int(aligned_risk_off.sum()),
                    "mixed_or_zero": int((~(aligned_risk_on | aligned_risk_off)).sum()),
                },
            },
            "temporal_semantics": {
                "release_month": "survey month encoded in the versioned official URL",
                "publication_date": "date printed in the release page header",
                "last_updated_date": "date printed in the page revision footer",
                "available_at": (
                    "23:59:59 Europe/Moscow on the later of publication_date and "
                    "last_updated_date"
                ),
                "admissible_join": "available_at less than or equal to decision_at",
                "date_only_source_uses_conservative_day_end": True,
                "release_specific_files_retrieved_currently": True,
                "original_historical_response_bytes_available": False,
                "historical_content_immutability_cryptographically_proved": False,
                "development_backtest_admissible": True,
                "independent_confirmation_without_forward_vintage_collection": False,
                "contains_prices_returns_targets_labels_or_pnl": False,
                "missing_values_are_not_zero": True,
            },
            "value_semantics": {
                "strategy_admissible_values": [
                    "exact XLSX median expected inflation over the next 12 months",
                    "exact XLSX consumer sentiment index",
                ],
                "observed_inflation_exact_retained_for_source_audit": True,
                "page_chart_endpoints_cross_checked_against_xlsx": True,
                "one_decimal_page_chart_display_retained": True,
                "latest_current_vintage_history_not_used": True,
            },
            "source_quality": {
                "archive_contains_every_expected_month": True,
                "archive_index_unchanged_during_collection": True,
                "every_release_has_page_pdf_and_xlsx": True,
                "every_xlsx_series_ends_on_release_month": True,
                "every_page_and_xlsx_inflation_display_endpoint_matches": True,
                "page_xlsx_exact_match_count": int(frame["page_xlsx_exact_match"].sum()),
                "duplicate_release_months": 0,
            },
            "rights": {
                "redistribution_license_verified": False,
                "raw_stored_outside_git": True,
            },
            "artifacts": {
                "processed": {
                    "path": data_path.name,
                    "bytes": data_path.stat().st_size,
                    "sha256": sha256_file(data_path),
                    "rows": len(frame),
                    "columns": frame.columns.tolist(),
                },
                "coverage": {
                    "path": coverage_path.name,
                    "bytes": coverage_path.stat().st_size,
                    "sha256": sha256_file(coverage_path),
                    "rows": len(coverage),
                },
                "raw_responses": {
                    "path": raw_path.name,
                    "bytes": raw_path.stat().st_size,
                    "sha256": sha256_file(raw_path),
                    "records": len(raw_records),
                },
            },
        }
        manifest_path = temporary / "manifest.json"
        write_json(
            manifest_path,
            {
                **manifest_core,
                "manifest_payload_sha256": sha256_bytes(_canonical_json(manifest_core)),
            },
        )
        manifest_sha = sha256_file(manifest_path)
        atomic_write_bytes(
            temporary / "manifest.sha256",
            f"{manifest_sha}  manifest.json\n".encode("utf-8-sig"),
        )
        temporary.replace(final)
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise
    return final


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Immutable external source-bundle directory.",
    )
    parser.add_argument("--max-workers", type=int, default=DEFAULT_MAX_WORKERS)
    arguments = parser.parse_args()
    print(
        download_cbr_inflation_expectations(
            arguments.output,
            max_workers=arguments.max_workers,
        )
    )


if __name__ == "__main__":
    main()
